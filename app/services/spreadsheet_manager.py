import os
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import locale
import pandas as pd
import re
import io
import base64


def processar_plan(plan, nome_orc, vend_orc, numero_orc, revisao_orc, novo_arquivo):
    # Abrindo a planilha de Cotação Distribuição
    filename2 = "data\modelos\MODELO COMPOSICAO DE VALORES - DISTRIBUIÇÃO 03.2025.xlsm"
    
    wb2 = openpyxl.load_workbook(filename2)
    ws2 = wb2.active
    
    planilha_ia = 'data\Itens IA\PLANILHA ITENS IA.xlsx'
    
    wbIA = openpyxl.load_workbook(planilha_ia)
    wsIA = wbIA.active
    
    #Contando quantos itens a Tabela tem
    IAr = wsIA.max_row
    IAc = wsIA.max_column
    IAa = IAr + 1  # Para corrigir ao adicionar itens. (pois a primeira linha é o nome da Coluna)

    print('O número de Linhas da Planilha Treinamento IA é', IAr) #N° de Linhas
    print('O número de Colunas da Planilha Treinamento IA é', IAc) #N° de Colunas
    
    
    
    

    # Colocando o Nome do Orçamentista e do Vendedor e o Fabricante
    ws2['C3'] = nome_orc
    ws2['C4'] = vend_orc
    ws2['C5'] = numero_orc
    ws2['C6'] = revisao_orc
    ws2['C8'] = 'TROX'


    # Abrindo a planilha do Tqs com os itens
    wb1 = openpyxl.load_workbook(plan)
    ws1 = wb1.worksheets[1]

    #Número de Linhas e colunas
    mr = ws1.max_row #N° de Linhas
    mc = ws1.max_column #N° de Colunas
    nl = mr + 1 #ultima linha de produtos
    start_row = 11 #Linha que começa os itens

    # Carregando a descrição e a descrição SKU
    values_qnt = [ws1.cell(row=i,column=17).value for i in range (2,nl)]
    values_copy = [ws1.cell(row=i,column=11).value for i in range (2,nl)] #Adicionando a descrição
    values_copy_sku = [ws1.cell(row=i,column=12).value for i in range (2,nl)] #Adicionando o código SKU

    #Adicionando Quantidade
    for i, value in enumerate(values_qnt):
        ws2.cell(row=i+11,column = 2).value = value

    # Adicionando a Descrição
    for i, value in enumerate(values_copy):
        value = str(value)
        value = '010 ' + value +  ' TROX'
        ws2.cell(row=i+11,column = 4).value = value

    # Adicionando o Grupo
    col_name_grupo = 'E'
    start_row = 11
    end_row = nl + 8
    col_name = 'C'
    for row in range(start_row, end_row + 1):
        cell = f'{col_name_grupo}{row}'
        ws2[cell] = 'C'


    # Adicionando o Código SKU
    for i, value in enumerate(values_copy_sku):
        ws2.cell(row=i+start_row,column = 7).value = value


    # Adicionando a Unidade
    for row in range(start_row, end_row + 1):
        cell = f'{col_name}{row}'
        ws2[cell] = 'PÇ'


    # Adicionando o NCM
    col_name_ncm = 'H'
    for i, value in enumerate(values_copy):
        if value in ('VENEZIANA','GRELHA DE ALETAS FIXAS','GRELHA DE ALETAS MOVEIS VERT.','DIFUSOR DE ALTA INDUCAO','GRELHA DE ALETAS MOVEIS HORIZ.','GRELHA INDEVASSAVEL','GRELHA PARA PISO','DIFUSOR MULTIVIAS QUADRADO','DIFUSOR MULTIVIAS RETANGULAR','DIFUSOR MULTIVIAS REDONDO','CAIXA PLENUM PARA GRELHAS','DIFUSOR LINEAR','CANTONEIRA TERMINAL ALS','TOMADA DE AR EXTERNO','CANTONEIRA TERMINAL VSD35','CANTONEIRA TERMINAL VSD50','DIFUSOR DE LONGO ALCANCE','GRELHA CONTINUA ALETAS FIXAS','GRELHA SEM MOLDURA','CANTONEIRA TERMINAL ADE'):
            NCM = 76169900
        elif value in ('DAMPER CORTA FOGO','ATENUADOR DE RUÍDO XS','ATENUADOR DE RUÍDO MS','PORTA DE INSPEÇÃO','ATENUADOR REDONDO','PORTA ACUSTICA'):
            NCM = 84798999
        elif value in ('DAMPER DE REGULAGEM','DAMPER DE REGULAGEM PESADO','DAMPER DE SOBRE PRESSAO','DAMPER DE ESTANQUEIDADE PESADO','REG. DE VAZAO CONSTANTE RET.','REG. DE VAZAO CONSTANTE RED.','CAIXA VAV RETANGULAR EASY','DAMPER REG MEDIA PRESSAO','CAIXA VAV REDONDA'):
            NCM = 90261029
        elif value in('FILTRO GROSSO FMP','FILTRO PLANO MFP','FILTRO MEDIO','FILTRO METALICO CORRUGADO','CAIXA TERMINAL ENTR. RED. SUP.','ESTRUTURA FILTRO PLANO PAREDE','FILTRO METALICO EM TELA','FILTRO BOLSA','ESTR. FILTRO METALICO PAREDE','ESTRUTURA PORTA FILTRO','CAIXA TERMINAL ENTR. RET. LAT.','CAIXA TERMINAL F670','FILTRO ABSOLUTO SEP. ALUM.','FILTRO CUNHA MFC','FILTRO MANTA'):
            NCM = 84213990
        elif value in ('CHAPA UNIAO - DIFUSOR ALS','CAIXA PLENUM AK6C','CAIXA PLENUM AK6'):
            NCM = 73269090
        elif value in ('PINO UNIAO - DIFUSOR ALS'):
            NCM = 73182400
        elif value in ('TERMOSTATO GS7.05.S GLOBUS'):
            NCM = 90328982
        elif value in ('CHAPA DE ALINHAMENTO VSD'):
            NCM = 72103010
        elif value in ('PINO DE INTERLIGACAO VSD'):
            NCM = 73182300
        else:
            NCM = 'COLOCAR NCM'

        ws2.cell(row=i+start_row,column = 8).value = NCM

    # Adicionando o Preço Unitário
    values_copy_preço = [ws1.cell(row=i,column=15).value for i in range (2,nl)]
    for i, value in enumerate(values_copy_preço):
        ws2.cell(row=i+start_row,column = 9).value = float(value)
        
        wsIA.cell(row=i+IAa,column = 27).value = float(value)


    # Adicionando Imposto ICMS ST
    values_copy_st = [ws1.cell(row=i,column=22).value for i in range (2,nl)]

    for i, value in enumerate(values_copy_st):
        ws2.cell(row=i+start_row,column = 10).value = value

    # Adicionando o Imposto IPI
    values_copy_ipi = [ws1.cell(row=i,column=21).value for i in range (2,nl)]

    for i, value in enumerate(values_copy_ipi):
        value = value/100
        ws2.cell(row=i+start_row,column = 11).value = value



    # Adicionando as medidas e acessórios

    for i, value in enumerate(values_copy):
        # Grelha AT (AT-0-AG125x125/00FAN0M0)
        if value in ('GRELHA DE ALETAS MOVEIS HORIZ.'):
            padrao_modelo = r"([A-Z]+)-"
            padrao_registro = r"-(AG|A|DG|D|Z)"
            padrao_dimensoes = r"(\d+)x(\d+)/"
            padrao_moldura = r"/(00|ERB11S|ERFB11S)"
            padrao_pintura = r"(AN0|PH1|PH2|PH3|PH4|PE0)"

            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            registro = re.search(padrao_registro, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio4 = 'SEM'


            if pintura == 'PH2':
                pintura = 'BRANCO'

            comprimento = int(comprimento)
            # Adicionando o Grupo
            if comprimento <= 325:
                grupo = 'A'
                ws2.cell(row=i+start_row,column = 5).value = grupo

            elif 325 < comprimento <= 525:
                grupo = 'B'
                ws2.cell(row=i+start_row,column = 5).value = grupo
            
            comprimento = str(comprimento)
            descricao = '010 ' + 'GRELHA' + ' ' + modelo +"-" +registro + " " + comprimento + "x" + largura + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 7).value = moldura
            wsIA.cell(row = i+IAa, column = 8).value = pintura

        # Grelha VAT - 	VAT-0-DG525x425/00FAN0M0
        if value in ('GRELHA DE ALETAS MOVEIS VERT.'):
            padrao_modelo = r"([A-Z]+)-"
            padrao_registro = r"-(AG|A|DG|D|Z)"
            padrao_dimensoes = r"(\d+)x(\d+)/"
            padrao_moldura = r"/(00|ERB11S|ERFB11S)"
            padrao_pintura = r"(AN0|PH1|PH2|PH3|PH4|PE0)"

            desc = values_copy_sku[i]
            desc_reg = desc[4:10]

            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            registro = re.search(padrao_registro, desc_reg).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio4 = 'SEM'

            if pintura == 'PH2':
                pintura = 'BRANCO'


            # Adicionando o Grupo
            comprimento = int(comprimento)
            if comprimento <= 325:
                grupo = 'A'
                ws2.cell(row=i+start_row,column = 5).value = grupo

            elif 325 < comprimento <= 525:
                grupo = 'B'
                ws2.cell(row=i+start_row,column = 5).value = grupo


            comprimento = str(comprimento)
            descricao = '010 ' + 'GRELHA' + ' ' + modelo + '-' + registro + ' ' + comprimento + "x" + largura + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 7).value = moldura
            wsIA.cell(row = i+IAa, column = 8).value = pintura

        if value in ('GRELHA DE ALETAS FIXAS'):

            padrao_modelo = r"([A-Z]+)"
            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            # Grelha AR -	AR-A-525x325/0/0/FAN000
            if modelo == 'AR':
                padrao_registro = r"-(AG|A|DG|D|Z)"
                padrao_dimensoes = r"(\d+)x(\d+)/"
                padrao_moldura = r"/(0|ER|ERF)"
                padrao_pintura = r"(AN0|PH1|PH2|PH3|PH4|PE0|PE4|PS3)"

                modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
                registro = re.search(padrao_registro, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0
                moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                acessorio4 = 'SEM'

                if pintura == 'PH2':
                    pintura = 'BRANCO'

                comprimento = int(comprimento)
                # Adicionando o Grupo
                if comprimento <= 325:
                    grupo = 'A'
                    ws2.cell(row=i+start_row,column = 5).value = grupo

                elif 325 < comprimento <= 525:
                    grupo = 'B'
                    ws2.cell(row=i+start_row,column = 5).value = grupo
                

                comprimento = str(comprimento)
                descricao = '010 ' + 'GRELHA' + ' ' + modelo + '-' + registro + ' ' + comprimento + "x" + largura + ' ' + pintura +" TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 3).value = registro
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 7).value = moldura
                wsIA.cell(row = i+IAa, column = 8).value = pintura

            if modelo in ('AH'):
                padrao_registro = r"-(AG|A|DG|D|Z)"
                padrao_dimensoes = r"(\d+)x(\d+(\.\d+)?)/"
                padrao_moldura = r"/(00|ERB11S|ERFB11S)"
                padrao_pintura = r"(AN0|PH1|PH2|PH3|PH4|PE0|PE4|PS3)"
                padrao_angulo = r"AH(\d+)-"

                modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
                registro = re.search(padrao_registro, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento = dimensoes.group(1)
                largura = dimensoes.group(2)
                #comprimento, largura = dimensoes.groups()
                espessura = 0
                moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                angulo = re.search(padrao_angulo, values_copy_sku[i]).group(1)

                if pintura == 'PH2':
                    pintura = 'BRANCO'

                comprimento = int(comprimento)
                # Adicionando o Grupo
                if comprimento <= 325:
                    grupo = 'A'
                    ws2.cell(row=i+start_row,column = 5).value = grupo

                elif 325 < comprimento <= 525:
                    grupo = 'B'
                    ws2.cell(row=i+start_row,column = 5).value = grupo
                
                
                comprimento = str(comprimento)
                descricao = '010 ' + 'GRELHA' + ' ' + modelo + '-'  + angulo +' ' + registro + ' ' + comprimento + "x" + largura + ' ' + pintura +" TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 3).value = registro
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 7).value = moldura
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 3).value = angulo

        # Grelha AGS - AGS-T/525x325/00/CF/AN0
        if value in ('GRELHA INDEVASSAVEL'):
            padrao_modelo = r"([A-Z]+(?:-[A-Z]+)?)\/"
            padrao_registro = r"-(AG|A|DG|D|Z)"
            padrao_dimensoes = r"(\d+)x(\d+)/"
            padrao_moldura = r"/(00|A1)"
            padrao_pintura = r"(AN0|PH1|PH2|PH3|PH4|PE0|PE4|PS3)"


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            registro = 'SEM'
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio4 = 'SEM'

            if pintura == 'PH2':
                pintura = 'BRANCO'

            comprimento = int(comprimento)
            # Adicionando o Grupo
            if comprimento <= 325:
                grupo = 'A'
                ws2.cell(row=i+start_row,column = 5).value = grupo

            elif 325 < comprimento <= 525:
                grupo = 'B'
                ws2.cell(row=i+start_row,column = 5).value = grupo
                
                
            comprimento = str(comprimento)
            descricao = '010 ' + 'GRELHA' + ' ' + modelo + ' ' + comprimento + "x" + largura + ' ' + pintura +" TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 7).value = moldura
            wsIA.cell(row = i+IAa, column = 8).value = pintura


        # Damper RG - RG-B-450x205/D/0/00/000
        if value in ('DAMPER DE REGULAGEM'):
            padrao_modelo = r"([A-Z]+(?:-[A-Z]+)?)-"
            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            if modelo in ('RG-B','RG-A'):
                padrao_dimensoes = r"-(\d+)x(\d+)/"
                padrao_lado_acio = r"/(D|E)/"           # Lado de acionamento
                padrao_tipo_acio = r"/(0|Z)/"           # Tipo de acionamento
                padrao_atuador = r"/(00|AA|AB|AC|AD|AE|AF|AG|AH|AI|AJ|AK|AL|AM|AN|AO|AP|AQ|AR|AS|AT)/"
                padrao_pintura = r"/(000|PEC|PE4)"

                modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0
                lado_acio = re.search(padrao_lado_acio, values_copy_sku[i]).group(1)
                tipo_acio = re.search(padrao_tipo_acio, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                atuador = re.search(padrao_atuador, values_copy_sku[i]).group(1)



                descricao = '010 ' + 'DAMPER' + ' ' + modelo + ' ' + comprimento + "x" + largura +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
            
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 10).value = tipo_acio
                wsIA.cell(row = i+IAa, column = 11).value = atuador
            

        # Damper KUL e KUL-CP e KUL-E - KUL/1025x500 - KUL-CP/500x500/000 - KUL-E/250x120
        if value in ('DAMPER DE SOBRE PRESSAO'):
            padrao_modelo = r"([A-Z]+(?:-[A-Z]+)?)"
            padrao_dimensoes = r"(\d+)x(\d+)"
            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)

            if modelo in ('KUL','KUL-CP','KUL-E'):

                registro = 'SEM'
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0
                moldura = 'SEM'
                pintura = 'SEM'
                acessorio4 = 'SEM'

                descricao = '010 ' + 'DAMPER' + ' ' + modelo + ' ' + comprimento + "x" + largura + " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao

                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura               
                


            # Damper UL - UL-2/200x350/CP/CF/PH2
            if modelo == 'UL':
                padrao_modeloul = r'-(\d)'
                padrao_dimensoes = r'/(\d+)x(\d+)/'
                padrao_contra_peso = r'/(CP|00)/'
                padrao_furos = r'/(CF|SF)/'
                padrao_pintura = r'/(PE0|PH1|PH2|PH4|000)'



                modeloul = re.search(padrao_modeloul, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                contra_peso = re.search(padrao_contra_peso, values_copy_sku[i]).group(1)
                furo = re.search(padrao_furos, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                espessura = 0
                acessorio4 = 'SEM'
                
                modelo = modelo + '-' + modeloul
                
                if pintura == 'PH2':
                    pintura = 'BRANCO'

                descricao = '010 ' + 'DAMPER' + ' ' + modelo + ' '+ contra_peso + ' ' + comprimento + "x" + largura + " "+ pintura +" TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 12).value = contra_peso
                

            #Damper AUL - AUL-2-CP/500x300/SF/000
            if modelo == 'AUL':
                padrao_modeloaul = r'AUL-(\d+(-CP)?)'
                registro = 'SEM'
                moldura = 'SEM'
                padrao_pintura = r"(000|PH1|PH2|PH3|PH4)"
                acessorio4 = 'SEM'

                modeloaul = re.search(padrao_modeloaul, values_copy_sku[i]).group(1)
                modeloaul = 'AUL-' + modeloaul
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0

                if pintura == 'PH2':
                    pintura = 'BRANCO'


                descricao = '010 ' + 'DAMPER' + ' ' + modeloaul + ' ' + comprimento + "x" + largura + " "+ pintura +" TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modeloaul
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 8).value = pintura

        # Filtro Grosso FMP - FMP-STD-G4/ENCP-PC/570x470x25
        if value in ('FILTRO GROSSO FMP'):
            padrao_modelo = r"([A-Z]+)-"
            padrao_versao = r"-(STD|BAC|GLA)-"
            padrao_dimensoes = r"/(\d+)x(\d+)x(\d+)"
            padrao_classificacao = r"-(G4|G3)/"
            padrao_moldura = r"/(ENCP-PC|ENCP-ZN|PAD-00|ENCZ-PC|ENCZ-ZN|ROL-00)/"

            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            versao = re.search(padrao_versao, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            classificacao = re.search(padrao_classificacao, values_copy_sku[i]).group(1)
            acessorio4 = 'SEM'
            grupo = 'B'


            descricao = '010 ' +'FILTRO' + ' ' + modelo + ' ' + versao +' ' + classificacao +' ' + comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            ws2.cell(row=i+start_row,column = 5).value = grupo

            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura
            wsIA.cell(row = i+IAa, column = 7).value = moldura
            wsIA.cell(row = i+IAa, column = 13).value = versao
            wsIA.cell(row = i+IAa, column = 14).value = classificacao


        # Filtro médio 	F754/ENCZ/600x500x48
        if value in ("FILTRO MEDIO"):
            padrao_modelo = r'([A-Z0-9]+)/'
            padrao_moldura = r"(ENCZ)"
            padrao_dimensoes = r"/(\d+)x(\d+)x(\d+)"

            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            classificacao = 'SEM'
            acessorio4 = 'SEM'
            versao = 'SEM'

            grupo = 'B'

            descricao = '010 ' +'FILTRO ' + modelo + ' ' + moldura +' ' + comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            ws2.cell(row=i+start_row,column = 5).value = grupo
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura
            wsIA.cell(row = i+IAa, column = 7).value = moldura



        # Filtro Metálico FMB - FMB-600x600x50-3240
        if value == 'FILTRO METALICO CORRUGADO':
            padrao_modelo = r'([A-Z]+)'
            padrao_dimensoes = r"-(\d+)x(\d+)x(\d+)-"
            acessório1 = 'SEM'
            acessório2 = 'SEM'
            acessório3 = 'SEM'
            acessório4 = 'SEM'


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(0)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()


            descricao = '010 ' + 'FILTRO '+ modelo + ' ' + comprimento + "x" + largura + "x" + espessura +  " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura



        # Estrutura de Filtro Metálico FMPA - FMPA/2440x2440/ZN
        if value in ('ESTR. FILTRO METALICO PAREDE'):
            padrao_modelo = r'([A-Z]+)'
            padrao_dimensoes = r"/(\d+)x(\d+)/"
            acessório1 = 'SEM'
            acessório2 = 'SEM'
            acessório3 = 'SEM'
            acessório4 = 'SEM'


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(0)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0



            descricao = '010 '+'EST.FILTRO ' + modelo + ' '+ comprimento + "x" + largura +  " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura


        #Tomada de ar Externo VDF - VDF-F711/297x297/0/0/SF/PH2
        if value == 'TOMADA DE AR EXTERNO':
            padrao_modelo = r"([A-Z]+)-"
            modelomed = values_copy_sku[i]
            modelomedcorte = modelomed[:16]
            venezianacorte = modelomed[16:]
            registrocorte = modelomed[18:]
            padrao_filtro = r"-([A-Z]+(\d+)?)/"
            padrao_dimensoes = r"/(\d+)x(\d+)/"
            padrao_veneziana = r"(K|G|0)"
            padrao_registro = r"(AG|0)"
            padrao_pintura = r"(PH1|PH2|PH3|PH4|AN0|PE0|PE4|PS3)"

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            filtro = re.search(padrao_filtro, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            veneziana = re.search(padrao_veneziana, venezianacorte).group(1)
            registro = re.search(padrao_registro, registrocorte).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)

            if pintura == 'PH2':
                pintura = 'BRANCO'

            elif pintura == 'PH4':
                pintura = 'PRETO'
                
            if veneziana == 'K':
                veneziana = 'AWK'
                
            elif veneziana == 'G':
                veneziana = 'AWG'
            
            elif veneziana == '0':
                veneziana = 'SEM VENEZIANA'


            descricao = '010 '+'TOMADA DE AR ' + modelo + ' '+ filtro + ' '+ veneziana + ' ' +comprimento + "x" + largura + ' '+  " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 15).value = filtro 
            wsIA.cell(row = i+IAa, column = 16).value = veneziana        


        #Atenuador de ruído XS e MS - XS-20/1200x900x900-4-100-1
        if value in ('ATENUADOR DE RUÍDO XS','ATENUADOR DE RUÍDO MS'):
            padrao_modelo = r"(XS-10|XS-20|MS-10|MS-20)/"
            padrao_dimensoes = r"/(\d+)x(\d+)x(\d+)"
            padrao_celulas = r"-(\d+)-(\d+(?:\.\d+)?)-(\d+)"

            modelo = re.search(padrao_modelo,values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            celulas = re.search(padrao_celulas, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            ncelula, tamcelula, acessoriocelula = celulas.groups()
            acessorio4 = 'SEM'


            descricao = '010 '+'ATENUADOR ' + modelo + ' ' + comprimento + "x" + largura + "x" + espessura +" " + ncelula + " " + tamcelula +" TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura
            wsIA.cell(row = i+IAa, column = 17).value = ncelula
            wsIA.cell(row = i+IAa, column = 18).value = tamcelula             


        # Damper JN - JN-B-M-E-N0/250x350/N/AD/I16
        if value == 'DAMPER DE REGULAGEM PESADO':
            print(value)

            padrao_modelo = r"([A-Z]+(?:-[A-Z]+)?)-"
            padrao_tipo_acio = r"-(0|M)-"
            padrao_lado_acio = r"(D|E)"
            padrao_vedacao = r"(N0|NV|L0|LV)"
            padrao_dimensoes = r"/(\d+)x(\d+)/"
            padrao_atuador = r"/(00|AA|AB|AC|AD|AE|AF|AG|AH|AI|AJ|AK|AL|AM|AN|AO|AP|AQ|AR|AS|AT)/"
            padrao_pintura = r"/(000|I16|I4|PE4|PEC|PH2)"


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            lado_acio = re.search(padrao_lado_acio, values_copy_sku[i]).group(1)
            tipo_acio = re.search(padrao_tipo_acio, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            atuador = re.search(padrao_atuador, values_copy_sku[i]).group(1)


            descricao = '010 '+'DAMPER ' + modelo + ' '+ tipo_acio + " " + comprimento + "x" + largura + ' ' + atuador +" "+ pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 10).value = tipo_acio
            wsIA.cell(row = i+IAa, column = 11).value = atuador


        # Veneziana AWG - AWG/875x330/N/MD/F/PH2
        if value == 'VENEZIANA':
            padrao_modelo = r"([A-Z]+)/"
            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            if modelo == "AWG":
                padrao_dimensoes = r"/(\d+)x(\d+)/"
                padrao_aleta = r"/(N|P)/"
                padrao_moldura = r"/(0|MD)/"
                padrao_pintura = r"/(AN0|PE4|PH2|PE0|PH4|PH3|PS3)"

                modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0
                aleta = re.search(padrao_aleta, values_copy_sku[i]).group(1)
                moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                acessorio4 = "SEM"

                if pintura == 'PH2':
                    pintura = 'BRANCO'

                elif pintura == 'PH4':
                    pintura = 'PRETO'

                descricao = '010 '+'VENEZIANA ' + modelo + " " + comprimento + "x" + largura + " " + pintura +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 7).value = moldura
                wsIA.cell(row = i+IAa, column = 8).value = pintura

            #Veneziana AWK - AWK/297x197/0/0/F/AN0/0
            if modelo == "AWK":
                padrao_dimensoes = r"/(\d+)x(\d+)/"
                padrao_fixacao = r"/(0|A11|B11)/"
                padrao_moldura = r"/(0|ER)/"
                padrao_furo_aba = r"(F|SF)"
                padrao_pintura = r"/(AN0|PE4|PH2|PE0|PH4|PH3|PS3|PH1)/"

                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                espessura = 0
                furo = re.search(padrao_furo_aba, values_copy_sku[i]).group(1)
                moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
                fixacao = re.search(padrao_fixacao, values_copy_sku[i]).group(1)

                if pintura == 'PH2':
                    pintura = 'BRANCO'

                elif pintura == 'PH4':
                    pintura = 'PRETO'


                descricao = '010 '+'VENEZIANA ' + modelo + " " + comprimento + "x" + largura + " " + pintura +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 5).value = largura
                wsIA.cell(row = i+IAa, column = 7).value = moldura
                wsIA.cell(row = i+IAa, column = 8).value = pintura


        #Difusor ADLQ - ADLQ-AA-0-0-T1/000SAN00M0 e ADLQ-A-0-0-T15P/000SAN0000
        if value == 'DIFUSOR MULTIVIAS QUADRADO':
            padrao_modelo = r"([A-Z]+)-"
            padrao_registro = r"-(AA|AB|AC|AD|AG|C|EA|QZ|A)-"
            padrao_caixa_plenum = r"-([0-3KVCK]+)-"
            padrao_fluxo = r"-(0|LDR)-"
            padrao_tamanho = r"-T([1-9])/"
            padrao_tamanho_polegada = r'-T(\d+)P/'
            padrao_pintura = r"(AN0|PH1|PH2|PE4|PH4)"

            tamanho_metro = re.search(padrao_tamanho,values_copy_sku[i])
            tamanho_polegada = re.search(padrao_tamanho_polegada,values_copy_sku[i])

            if tamanho_metro:
                tamanho  = tamanho_metro.group(1)
                unidade = 'MM'
            elif tamanho_polegada:
                tamanho  = tamanho_polegada.group(1)
                unidade = 'POL'
            else:
                tamanho = 'nada'
                unidade = None

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            registro = re.search(padrao_registro, values_copy_sku[i]).group(1)
            caixaplenum = re.search(padrao_caixa_plenum, values_copy_sku[i]).group(1)
            fluxo = re.search(padrao_fluxo, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)


            tamanho = int(tamanho)

            if tamanho < 6:
                grupo = 'B'
                ws2.cell(row=i+start_row,column = 5).value = grupo


            descricao = '010 '+'DIFUSOR ' + modelo + ' '+ "T" + str(tamanho) + unidade + " " + registro + " " + caixaplenum + " " + pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 19).value = caixaplenum
            wsIA.cell(row = i+IAa, column = 8).value = str(tamanho)



        # Difusor ADQ - ADQ-32KAG471x208F123SAN00M00 and ADQ-30AGT21PxT9PF000SAN00M00
        if value == 'DIFUSOR MULTIVIAS RETANGULAR':

            padrao_modelo = r"ADQ-[1-4][A-Z0-9]?"
            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(0)
            info = values_copy_sku[i][3:]                                     #Foi feito para que o código não encontra o A de ADQ como registro.
            padrao_caixa_plenum = r"(K|0)"
            padrao_registro = r"(AG|A)"
            padrao_dimensoes = r"(\d+)x(\d+)"
            padrao_pintura = r"(AN0|PH1|PH2|PE4|PH4)"
            padrao_miolo = r"(R)*"
            padrao_dimensoes_polegadas = r"T(\d+)PxT(\d+)P"

            dimensoes_metros = re.search(padrao_dimensoes, values_copy_sku[i])
            dimensoes_polegadas = re.search(padrao_dimensoes_polegadas, values_copy_sku[i])

            if dimensoes_metros:
                comprimento, largura = dimensoes_metros.groups()
                unidade = 'MM'
            elif dimensoes_polegadas:
                comprimento, largura = dimensoes_polegadas.groups()
                unidade = 'POL'
            else:
                comprimento, largura = None, None
                unidade = None


            dimensoes_polegadas = re.search(padrao_dimensoes_polegadas, values_copy_sku[i])
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])

            caixaplenum = re.search(padrao_caixa_plenum, info).group(0)
            registro = re.search(padrao_registro, info).group(0)
            espessura = 0
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(0)
            miolo = re.search(padrao_miolo, values_copy_sku[i]).group(0)



            if modelo in ('ADQ-10','ADQ-20','ADQ-30','ADQ-40','ADQ-1K','ADQ-2K','ADQ-3K','ADQ-4K'): #Feito para que quando o modelo for ADQ-1 e não ADQ-1C ele pegue o modelo certo
                modelo = modelo[:5]

            if dimensoes_metros:
                descricao = '010 '+'DIFUSOR ' + modelo + ' '+ comprimento + "x" + largura +' '+ unidade +" " + registro + " " + caixaplenum + " " + pintura +   " TROX"
            elif dimensoes_polegadas:
                descricao = '010 '+'DIFUSOR ' + modelo + ' '+ comprimento + "'x" + largura + "' " + unidade +" " + registro + " " + caixaplenum + " " + pintura +   " TROX"
            else:
                comprimento, largura = None, None
                unidade = None

            
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 19).value = caixaplenum


        #Filtro Médio F74B - F74B33/I4/287x592x600-2150
        if value == 'FILTRO BOLSA':
            padrao_modelo = r'([A-Z0-9]+[A-Z]?(\d+)?)'
            padrao_moldura = r'/(M|I4)/'
            padrao_dimensoes = r'/(\d+)x(\d+)x(\d+)-'



            modelotudo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            modelo = modelotudo[:3]
            filtro = modelotudo[3:]
            moldura = re.search(padrao_moldura, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            acessório3 = 'SEM'
            acessorio4 = 'SEM'
            if filtro == "B31":
                filtro = 'M6'
            elif filtro == "B33":
                filtro = 'F9'
            elif filtro == "BSB65":
                filtro = 'M6'
            elif filtro == "BSB85":
                filtro = 'F7'
            elif filtro == "BSB95":
                filtro = 'F8'


            descricao = '010 '+'FILTRO ' + modelo + ' ' + filtro +' ' + comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura
            wsIA.cell(row = i+IAa, column = 7).value = moldura
            wsIA.cell(row = i+IAa, column = 14).value = filtro
            
            


        # Filtro Plano MFP - MFP-H13-ALZ/305x305x78(85)x45/00/FND/OTC/0 - MFP-ePM1-90%-ALZ/305x305x78(85)x55/00/FND/000/0
        if value == 'FILTRO PLANO MFP':

            padrao_filtro = r'(H13|H14|ePM10|ePM1)'
            padrao_const = r'-(ALL|ALSX|ALS|ALUX|ALU|ALVX|ALV|ALZX|ALZ|GALF|GALP|GAL|MDF|STA)/'
            padrao_dimensoes = r"/(\d+)x(\d+)x(\d+)"
            padrao_vedacao = r"/(FNB|FND|FNU|GPU|WS)/"
            padrao_tela = r'(00|PB|PD|PU|OCT)'


            modelo = 'MFP'
            filtro = re.search(padrao_filtro, values_copy_sku[i]).group(1)
            const = re.search(padrao_const, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            vedacao = re.search(padrao_vedacao, values_copy_sku[i]).group(1)
            tela = re.search(padrao_tela, values_copy_sku[i]).group(1)
            print(filtro)

            if filtro == 'ePM10':
                filtro = 'M6'


            elif filtro == 'ePM1':
                f_plano = values_copy_sku[i]
                filtro_porcent = f_plano[9:11]
                if filtro_porcent == '60':
                    filtro = 'F7'
                if filtro_porcent == '90':
                    filtro = 'F9'


            descricao = '010 '+ 'FILTRO ' + modelo + ' ' + filtro +' ' + const +' ' + vedacao+ ' ' + comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura
            wsIA.cell(row = i+IAa, column = 14).value = filtro
            wsIA.cell(row = i+IAa, column = 21).value = const
            wsIA.cell(row = i+IAa, column = 22).value = vedacao         
            
            


        # Filtro Metálico em Tela F716 - F716-250x400x18/900
        if value == 'FILTRO METALICO EM TELA':
            padrao_modelo = r'([A-Z0-9]+)'
            padrao_dimensoes = r'-(\d+)x(\d+)x(\d+)/'

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            acessório1 = "SEM"
            acessório2 = "SEM"
            acessório3 = "SEM"
            acessório4 = "SEM"


            descricao = '010 '+'FILTRO ' + modelo +' ' +comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao

            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura

        # Porta de Inspeção IT - IT-25/500x400
        if value == 'PORTA DE INSPEÇÃO':

            padrao_modelo = r'([A-Z]+)'
            padrao_dimensoes = r"/(\d+)x(\d+)"
            acessório1 = 'SEM'
            acessório2 = 'SEM'
            acessório3 = 'SEM'
            acessório4 = 'SEM'


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(0)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0


            descricao = '010 ' + 'PORTA DE INSPEÇÃO ' + modelo + ' '+ comprimento + "x" + largura +  " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura


        # Filtro Absoluto F77 - F771E610x610x292CN2000E#29200
        if value == 'FILTRO ABSOLUTO SEP. ALUM.':
            # padrao_modelo = r'([A-Z0-9]+)'
            padrao_modelo = r'([A-Z]+[0-9]+)'
            padrao_dimensoes = r'(\d+)x(\d+)x(\d+)'
            padrao_carcaca = r'(E|K|M)'
            padrao_construcao = r'(CN|HT)'
            padrao_vedação = r'(D|E|S)'

            desc = values_copy_sku[i]
            desc_1 = desc[:16]
            desc_2 = desc[16:]


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura, espessura = dimensoes.groups()
            carcaca = re.search(padrao_carcaca, desc_1).group(1)
            construcao = re.search(padrao_construcao, values_copy_sku[i]).group(1)
            vedacao = re.search(padrao_modelo, desc_2).group(1)
            acessório4 = "SEM"


            descricao = '010 '+ 'FILTRO ' + modelo + ' ' + comprimento + "x" + largura + "x" + espessura + " MM" + " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 6).value = espessura


        #Damper JN-ATF - JN-ATF-A-MDN/500x510N/AE/PE4
        if value == 'DAMPER DE ESTANQUEIDADE PESADO':

            padrao_modelo = r"(([A-Z]+)-([A-Z]+)?-([A-Z]+)?)"
            padrao_tipo_acio = r"(0|M)"
            padrao_lado_acio = r"(D|E)"
            padrao_dimensoes = r"/(\d+)x(\d+)"
            padrao_atuador = r"/(00|AA|AB|AC|AD|AE|AF|AG|AH|AI|AJ|AK|AL|AM|AN|AO|AP|AQ|AR|AS|AT)/"
            padrao_pintura = r"/(000|PE4|PEC)"


            desc = values_copy_sku[i]
            desc_acio = desc[:12]
            desc_pintu = desc[23:]


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)

            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            lado_acio = re.search(padrao_lado_acio, values_copy_sku[i]).group(1)
            tipo_acio = re.search(padrao_tipo_acio, desc_acio).group(1)
            pintura = re.search(padrao_pintura, desc_pintu).group(1)
            atuador = re.search(padrao_atuador, values_copy_sku[i]).group(1)


            descricao = '010 '+'DAMPER ' + modelo + ' '+ tipo_acio + " " + comprimento + "x" + largura + ' ' + atuador +" "+ pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 10).value = tipo_acio
            wsIA.cell(row = i+IAa, column = 11).value = atuador


        # Grelha para piso AF - AF-15-0-DG/500x300/PE4/0/M
        if value == 'GRELHA PARA PISO':
            padrao_modelo = r"(AF-0|AF-15|AF-C-0|AF-C-15)-"
            padrao_registro = r'(AG|A|DG|D)+'
            padrao_dimensoes = r"/(\d+)x(\d+)/"
            padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"

            desc = values_copy_sku[i]
            desc_reg = desc[5:]


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            registro = re.search(padrao_registro, desc_reg).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio2 = 'SEM'
            acessorio4 = 'SEM'

            comprimento = int(comprimento)
            # Adicionando o Grupo
            if comprimento <= 325:
                grupo = 'A'
                ws2.cell(row=i+start_row,column = 5).value = grupo

            elif 325 < comprimento <= 525:
                grupo = 'B'
                ws2.cell(row=i+start_row,column = 5).value = grupo
            
            
            comprimento = str(comprimento)
            descricao = '010 '+'GRELHA ' + modelo + ' '+ registro + " " + comprimento + "x" + largura +" "+ pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 8).value = pintura    



        # Difusor ALS - ALS-S3-2000/K/B/218/0/PH2
        if value == 'DIFUSOR LINEAR':
            padrao_modelo = r'([A-Z]+)'
            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)

            if modelo == 'ALS':

                padrao_aberturas = r'-(DS1|DS2|DS3|DS4|S1|S2|S3|S4)-'
                padrao_comprimento = r'-(\d+)/'
                padrao_caixa_plenum = r'/(0|K)/'
                padrao_cabeceiras = r'/(A|B|C|D)/'
                padrao_reg_col = r'/(0|M)/'
                padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"



                aberturas = re.search(padrao_aberturas, values_copy_sku[i]).group(1)
                comprimento = re.search(padrao_comprimento, values_copy_sku[i]).group(1)
                largura = 0
                espessura = 0
                caixa_plenum = re.search(padrao_caixa_plenum, values_copy_sku[i]).group(1)
                cabeceira = re.search(padrao_cabeceiras, values_copy_sku[i]).group(1)
                reg_col = re.search(padrao_reg_col, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)


                if caixa_plenum == 'K':
                    caixa_plenum = 'C/ CAIXA PLENUM'
                    descricao = '010 '+'DIFUSOR ' + modelo + ' ' + aberturas + " " + comprimento + " " + caixa_plenum + " " + pintura +   " TROX"

                elif caixa_plenum == '0':
                    descricao = '010 '+'DIFUSOR ' + modelo + ' ' + aberturas + " " + comprimento + " " + pintura +   " TROX"



                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 19).value = caixa_plenum 
                wsIA.cell(row = i+IAa, column = 24).value = aberturas
                wsIA.cell(row = i+IAa, column = 25).value = cabeceira


            # Difusor linear VSD - VSD35-1F02000x123x0000C1AN0
            if modelo == 'VSD':
                padrao_n_modelo = r'VSD(\d+)'
                padrao_aberturas = r'-(\d)'
                padrao_comprimento = r'(\d+)x'
                padrao_caixa_plenum = r'(AA|AK|AS|DK|DS|F|VS)'
                padrao_cabeceiras = r'(A1|A2|A3|A4|A5|A6|B1|B2|B3|B4|B5|B6|C1|C2|C3|C4|C5|C6|0)'
                padrao_reg_col = r'(M|0)'
                padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"

                desc = values_copy_sku[i]
                desc_cabe = desc[-5:]
                print(desc_cabe)
                desc_caixa = desc[5:12]

                n_modelo = re.search(padrao_n_modelo, values_copy_sku[i]).group(1)
                aberturas = re.search(padrao_aberturas, values_copy_sku[i]).group(1)
                comprimento = re.search(padrao_comprimento, values_copy_sku[i]).group(1)
                comprimento = int(comprimento)
                comprimento = str(comprimento)
                largura = 0
                espessura = 0
                caixa_plenum = re.search(padrao_caixa_plenum, desc_caixa).group(1)
                cabeceira = re.search(padrao_cabeceiras, desc_cabe).group(1)
                reg_col = re.search(padrao_reg_col, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)


                descricao = '010 '+'DIFUSOR ' + modelo + ' ' + n_modelo + ' '+ aberturas + " " + comprimento + " " + cabeceira + " " + caixa_plenum + " " + pintura +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo + n_modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 19).value = caixa_plenum 
                wsIA.cell(row = i+IAa, column = 24).value = aberturas
                wsIA.cell(row = i+IAa, column = 25).value = cabeceira


        # Cantoneira ALS - ALS-CANT-1/AN0
        if value == 'CANTONEIRA TERMINAL ALS':
            padrao_modelo = r'(ALS-CANT)'
            padrao_aberturas = r'-(\d)/'
            padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            comprimento = 0
            largura = 0
            espessura = 0
            aberturas = re.search(padrao_aberturas, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio2 = 'SEM'
            acessorio4 = 'SEM'



            descricao = '010 '+'CANTONEIRA ' + modelo + ' ' + aberturas + " " + pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 24).value = aberturas



        # Cantoneira VSD - VSD35-CANT-2/CTT/000/PE0
        if value in ('CANTONEIRA TERMINAL VSD35','CANTONEIRA TERMINAL VSD50'):
            padrao_modelo = r'([A-Z]+[0-9]+)-'
            padrao_cant = r'/(CHT|CTT)/'
            padrao_aberturas = r'-(\d)/'
            padrao_perfil = r'/(000|A00|B00)/'
            padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            cantoneira = re.search(padrao_cant, values_copy_sku[i]).group(1)
            comprimento = 0
            largura = 0
            espessura = 0
            aberturas = re.search(padrao_aberturas, values_copy_sku[i]).group(1)
            perfil = re.search(padrao_perfil, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio4 = 'SEM'


            descricao = '010 '+'CANTONEIRA ' + modelo + ' ' + cantoneira + ' ' + aberturas + " " + pintura + ' ' + perfil +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo + cantoneira
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 24).value = aberturas
            wsIA.cell(row = i+IAa, column = 27).value = perfil
            



        #Damper JN-MP - JN-MP-B-MDN0-500x350-00000CF
        if value == 'DAMPER REG MEDIA PRESSAO':

            padrao_modelo = r"(([A-Z]+)-([A-Z]+)?-([A-Z]+)?)"
            padrao_tipo_acio = r"(0|M)"
            padrao_lado_acio = r"(D|E)"
            padrao_dimensoes = r"-(\d+)x(\d+)-"
            padrao_atuador = r"-(00|AA|AB|AC|AD|AE|AF|AG|AH|AI|AJ|AK|AL|AM|AN|AO|AP|AQ|AR|AS|AT)"
            padrao_pintura = r"(000|PE4|PEC)"


            desc = values_copy_sku[i]
            desc_tipo = desc[7:]
            desc_acio = desc[:12]
            desc_pintu = desc[23:]


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)

            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            lado_acio = re.search(padrao_lado_acio, values_copy_sku[i]).group(1)
            tipo_acio = re.search(padrao_tipo_acio, desc_tipo).group(1)
            pintura = re.search(padrao_pintura, desc_pintu).group(1)
            atuador = re.search(padrao_atuador, values_copy_sku[i]).group(1)


            descricao = '010 '+'DAMPER ' + modelo + ' '+ tipo_acio + " " + comprimento + "x" + largura + ' ' + atuador +" "+ pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 10).value = tipo_acio
            wsIA.cell(row = i+IAa, column = 11).value = atuador         


        # Regulador de vazão EN - EN/500x400/00
        if value == 'REG. DE VAZAO CONSTANTE RET.':
            padrao_modelo = r'([A-Z]+)/'
            padrao_dimensoes = r'/(\d+)x(\d+)/'

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            acessorio1 = 'SEM'
            acessorio2 = 'SEM'
            acessorio3 = 'SEM'
            acessorio4 = 'SEM'


            descricao = '010 ' + 'REG. VAZÃO CONST. ' + modelo + " " + comprimento + "x" + largura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura



        # Regulador de vazão redondo RN - RN/T100/00/00
        if value == 'REG. DE VAZAO CONSTANTE RED.':
            padrao_modelo = r'([A-Z]+)/'
            padrao_dimensoes = r'/T(\d+)/'

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            comprimento = re.search(padrao_dimensoes, values_copy_sku[i]).group(1)
            largura= 0
            espessura = 0
            acessorio1 = 'SEM'
            acessorio2 = 'SEM'
            acessorio3 = 'SEM'
            acessorio4 = 'SEM'


            descricao = '010 ' + 'REG. VAZÃO CONST. RED. ' + modelo + " " +"T" + comprimento +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento



        # Estrutura para filtro plano FPPA - FPPA-1223x1525/I04/N
        if value == 'ESTRUTURA FILTRO PLANO PAREDE':
            padrao_modelo = r'([A-Z]+)'
            padrao_dimensoes = r"-(\d+)x(\d+)/"
            padrao_material = r"(ZN|I04|I16)"


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(0)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            espessura = 0
            material = re.search(padrao_material, values_copy_sku[i]).group(0)
            acessório2 = 'SEM'
            acessório3 = 'SEM'
            acessório4 = 'SEM'


            descricao = '010 '+'EST. FILTRO ' + modelo + ' '+ comprimento + "x" + largura +  " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura
            


        # Difusor de Alta Indução VD - VD-H-0-K-0/1050/AF/AN0/0/0
        if value == 'DIFUSOR DE ALTA INDUCAO':
            padrao_modelo = r"([A-Z]+)"
            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)

            if modelo == 'VD':
                padrao_dimensoes = r'/(\d+)/'
                padrao_caixa_plenum = '-(0|HM|H|VM|V)-'
                padrao_placa = r"(K|0)"
                padrao_protecao = r'(S|0)'
                padrao_montagem = r'/(00|AF|RF)/'
                padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0)"
                padrao_espuma = r'/(0|1)'

                desc = values_copy_sku[i]
                desc_placa = desc[6:8]
                desc_prot = desc[8:10]
                print(desc_placa)

                modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
                comprimento = re.search(padrao_dimensoes, values_copy_sku[i]).group(1)
                largura= 0
                espessura = 0
                caixa_plenum = re.search(padrao_caixa_plenum, values_copy_sku[i]).group(1)
                placa = re.search(padrao_placa, desc_placa).group(0)
                protecao = re.search(padrao_protecao, desc_prot).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(0)


                descricao = '010 '+ 'DIFUSOR ' + modelo + ' '+ caixa_plenum + " " + comprimento +" "+ pintura + " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 19).value = caixa_plenum


                if caixa_plenum == "0":
                    descricao = '010 ' + modelo + " " + comprimento +" "+ pintura + " TROX"
                    
                    


            # Difusor de Alta Indução FD - FD-R-Z-V/400/158/M/PH1
            if modelo == 'FD':
                padrao_dimensoes = r'/(\d+)/'
                padrao_difusor = r'-(R|Q)-'
                padrao_caixa_plenum = '(0|H|V)'
                padrao_insuflamento = r"(A|Z)"
                padrao_pintura = r"(AN0|PH1|PH2|PH4|PS3|PE4|PE0|PP5)"

                desc = values_copy_sku[i]
                desc_plenum = desc[:8]

                comprimento = re.search(padrao_dimensoes, values_copy_sku[i]).group(1)
                largura= 0
                espessura = 0
                caixa_plenum = re.search(padrao_caixa_plenum, desc_plenum).group(1)
                insuflamento = re.search(padrao_insuflamento, values_copy_sku[i]).group(0)
                difusor = re.search(padrao_difusor, values_copy_sku[i]).group(1)
                pintura = re.search(padrao_pintura, values_copy_sku[i]).group(0)


                descricao = '010 '+ 'DIFUSOR ' + modelo + ' ' + " " + comprimento + " " + pintura + " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao
                
                wsIA.cell(row = i+IAa, column = 2).value = modelo
                wsIA.cell(row = i+IAa, column = 4).value = comprimento
                wsIA.cell(row = i+IAa, column = 8).value = pintura
                wsIA.cell(row = i+IAa, column = 19).value = caixa_plenum


                if caixa_plenum == "0":
                    descricao = '010 ' + modelo + " " + caixa_plenum + " " + comprimento + " "+ pintura + " TROX"
                    ws2.cell(row=i+start_row,column = 4).value = descricao


        # Difusor Redondo ADLR - ADLR-A-000-0/T1/0/PP5
        if value == 'DIFUSOR MULTIVIAS REDONDO':
            padrao_modelo = r"([A-Z]+)-"
            padrao_registro = r"-(A|C|KS|SZR|TG|ZH|ZV)-"
            padrao_fluxo = r"-(000|LDR)-"
            padrao_tamanho = r"/T([1-9])/"
            padrao_pintura = r"/(PH1|PH2|PE4|SP|PP5|PE0|PH4)"


            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            registro = re.search(padrao_registro, values_copy_sku[i]).group(1)
            acessorio2= 'SEM'
            fluxo = re.search(padrao_fluxo, values_copy_sku[i]).group(1)
            tamanho = re.search(padrao_tamanho,values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)


            descricao = '010 '+'DIFUSOR ' + modelo + ' '+ "T" + tamanho + " " + registro + " " + pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 3).value = registro
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 20).value = tamanho


        # DIFUSOR DE LONGO ALCANCE DUK - DUK-V-0-00/315/PH4
        if value == "DIFUSOR DE LONGO ALCANCE":
            padrao_modelo = r'([A-Z]+)'
            padrao_tamanho = r'/(\d+)/'
            padrao_pintura = r"/(PH1|PH2|PH5|PH4)"

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            tamanho = re.search(padrao_tamanho, values_copy_sku[i]).group(1)
            pintura = re.search(padrao_pintura, values_copy_sku[i]).group(1)
            acessorio1 = 'SEM'
            acessorio2 = 'SEM'
            acessorio4 = 'SEM'


            descricao = '010 '+'DIFUSOR ' + modelo + ' '+ "T" + tamanho + " " + pintura +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 8).value = pintura
            wsIA.cell(row = i+IAa, column = 20).value = tamanho            


        # CAIXA VAV RETANGULAR TVJ-TVT- EASY - TVJDEASY/1000x307
        if value == 'CAIXA VAV RETANGULAR EASY':
            padrao_modelo = r'([A-Z]+)/'
            padrao_dimensoes = r'/(\d+)x(\d+)'

            modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
            dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
            comprimento, largura = dimensoes.groups()
            acessorio1 = 'SEM'
            acessorio2 = 'SEM'
            acessorio3 = 'SEM'
            acessorio4 = 'SEM'


            descricao = '010 '+'CAIXA VAV ' + modelo + ' ' + comprimento + "x" + largura + " " +   " TROX"
            ws2.cell(row=i+start_row,column = 4).value = descricao
            
            
            wsIA.cell(row = i+IAa, column = 2).value = modelo
            wsIA.cell(row = i+IAa, column = 4).value = comprimento
            wsIA.cell(row = i+IAa, column = 5).value = largura



        # Damper corta fogo FKA - FKATIBR120-2-500x450-8-X00
        if value == 'DAMPER CORTA FOGO':
            
            padrao_modelo_ti = r'(FKATIBR)'
            modelo_ti = re.search(padrao_modelo_ti, values_copy_sku[i])


            padrao_modelo_nb = r'(NB)'
            modelo_nb = re.search(padrao_modelo_nb, values_copy_sku[i])


            padrao_modelo_ul = r'(FKAUL)'
            modelo_ul = re.search(padrao_modelo_ul, values_copy_sku[i])
        

            if modelo_ti:

                padrao_modelo = r'([A-Z]+[0-9]+)-'
                padrao_acabamento = r'-(\d+)-'
                padrao_dimensoes = r'-(\d+)x(\d+)-'
                padrao_prolongamento = r'-(\d+)-'
                padrao_fusivel = r'(V17|V17.1|X00|X01|X02|X17|X17.1|Y01|Y02|Y03|Y21|Y22|Y24|Y31|Y32|Y33|Y34|Y35|Y36|Z00|Z01|Z02|Z03|Z11|Z12|Z13|Z14|Z15|Z15.1|Z16|Z16.1|Z17|Z17.1|Z18|Z22)'

                desc = values_copy_sku[i]
                desc_meio1 = desc[:18]
                desc_meio2 = desc[18:]

                modelo = re.search(padrao_modelo, values_copy_sku[i]).group(1)
                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                acabamento = re.search(padrao_acabamento, desc_meio1).group(1)
                prolongamento = re.search(padrao_prolongamento, desc_meio2).group(1)
                acessorio3 = 'SEM'
                fusivel = re.search(padrao_fusivel, values_copy_sku[i]).group(1)

                if modelo == 'FKATIBR120':
                    modelo = 'FKA TI BR 120'

                if modelo == 'FKATIBR60':
                    modelo = 'FKA TI BR 60'

                if modelo == 'FKATIBR90':
                    modelo = 'FKA TI BR 90'

                descricao = '010 '+'DAMPER CORTA FOGO ' + modelo + ' ' + comprimento + "x" + largura+ 'MM' + " " + fusivel +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao


            # Corta Fogo FKA-NB-BR - FKA-NB-BR-0/500x500/0/Y01
            elif modelo_nb:
                modelo_nb = modelo_nb.group(0)
                print(modelo_nb)

                padrao_acabamento = r'-(\d+)/'
                padrao_dimensoes = r'/(\d+)x(\d+)/'
                padrao_prolongamento = r'/(\d+)/'
                padrao_fusivel = r'(V17|V17.1|X00|X01|X02|X17|X17.1|Y01|Y02|Y03|Y21|Y22|Y24|Y31|Y32|Y33|Y34|Y35|Y36|Z00|Z01|Z02|Z03|Z11|Z12|Z13|Z14|Z15|Z15.1|Z16|Z16.1|Z17|Z17.1|Z18|Z22)'

                desc = values_copy_sku[i]
                desc_meio1 = desc[:18]
                desc_meio2 = desc[18:]

                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                acabamento = re.search(padrao_acabamento, desc_meio1).group(1)
                prolongamento = re.search(padrao_prolongamento, desc_meio2).group(1)
                acessorio3 = 'SEM'
                fusivel = re.search(padrao_fusivel, values_copy_sku[i]).group(1)


                descricao = '010 '+'DAMPER CORTA FOGO FKA-NB-BR '  + comprimento + "x" + largura+ 'MM' + " " + fusivel +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao

            # Damper Corta Fogo FKAUL - FKAUL55590-1/500x500/0/X01
            elif modelo_ul:
                modelo_ul = modelo_ul.group(0)
                print(modelo_ul)

                padrao_acabamento = r'-(\d+)/'
                padrao_dimensoes = r'/(\d+)x(\d+)/'
                padrao_prolongamento = r'/(\d+)/'
                padrao_fusivel = r'(V17|V17.1|X00|X01|X02|X17|X17.1|Y01|Y02|Y03|Y21|Y22|Y24|Y31|Y32|Y33|Y34|Y35|Y36|Z00|Z01|Z02|Z03|Z11|Z12|Z13|Z14|Z15|Z15.1|Z16|Z16.1|Z17|Z17.1|Z18|Z22)'

                desc = values_copy_sku[i]
                desc_meio1 = desc[:18]
                desc_meio2 = desc[18:]

                dimensoes = re.search(padrao_dimensoes, values_copy_sku[i])
                comprimento, largura = dimensoes.groups()
                acabamento = re.search(padrao_acabamento, desc_meio1).group(1)
                prolongamento = re.search(padrao_prolongamento, desc_meio2).group(1)
                acessorio3 = 'SEM'
                fusivel = re.search(padrao_fusivel, values_copy_sku[i]).group(1)

                descricao = '010 '+'DAMPER CORTA FOGO FKA-UL-555 '  + comprimento + "x" + largura+ 'MM' + " " + fusivel +   " TROX"
                ws2.cell(row=i+start_row,column = 4).value = descricao


    # Caminho para salvar a nova planilha
    wb2.save(novo_arquivo)
    wbIA.save(planilha_ia)

