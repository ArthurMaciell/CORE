from datetime import datetime
import openpyxl

from app.utils.tempo_util import calcular_horas_uteis


def cadastrar_orcamento(planilha_cadastro_path: str, dados: dict):
    """
    Cadastra um novo orçamento na planilha de controle.

    Parâmetros:
    - planilha_cadastro_path: caminho da planilha .xlsx
    - dados: dicionário com todos os campos necessários
    """
    try:
        wb = openpyxl.load_workbook(planilha_cadastro_path)
        sheet = wb.worksheets[0]

        hora_atual_plan = datetime.now().strftime('%H:%M')

        horas_uteis_plan = calcular_horas_uteis(
            dados["data_orc_ini"],
            dados["hora_orc_ini"],
            dados["data_orc_conc"],
            hora_atual_plan
        )

        sheet_r = sheet.max_row
        sheet_a = sheet_r + 1

        sheet.cell(row=sheet_a, column=1).value = dados["data_orc_ini"]
        sheet.cell(row=sheet_a, column=2).value = dados["hora_orc_ini"]
        sheet.cell(row=sheet_a, column=3).value = dados["data_orc_conc"]
        sheet.cell(row=sheet_a, column=4).value = hora_atual_plan
        sheet.cell(row=sheet_a, column=5).value = dados["mes_por_extenso_ini"]
        sheet.cell(row=sheet_a, column=6).value = dados["dia_semana_ini"]
        sheet.cell(row=sheet_a, column=7).value = dados["mes_por_extenso"]
        sheet.cell(row=sheet_a, column=8).value = dados["dia_semana"]
        sheet.cell(row=sheet_a, column=9).value = dados["nome_orc"]
        sheet.cell(row=sheet_a, column=10).value = dados["fabrica_orc"]
        sheet.cell(row=sheet_a, column=11).value = dados["terceiros_orc"]
        sheet.cell(row=sheet_a, column=12).value = dados["frete_orc"]
        sheet.cell(row=sheet_a, column=13).value = dados["tipo_orc"]
        sheet.cell(row=sheet_a, column=14).value = dados["loja_orc"]
        sheet.cell(row=sheet_a, column=15).value = dados["vend_orc"]
        sheet.cell(row=sheet_a, column=16).value = dados["seguimento_orc"]
        sheet.cell(row=sheet_a, column=17).value = dados["cliente_orc"]
        sheet.cell(row=sheet_a, column=18).value = dados["icms_orc"]
        sheet.cell(row=sheet_a, column=19).value = dados["obra_orc"]
        sheet.cell(row=sheet_a, column=20).value = dados["local_obra"]
        sheet.cell(row=sheet_a, column=21).value = dados["tamanho_orc"]
        sheet.cell(row=sheet_a, column=22).value = dados["numero_orc"]
        sheet.cell(row=sheet_a, column=23).value = dados["revisao_orc"]
        sheet.cell(row=sheet_a, column=24).value = horas_uteis_plan
        sheet.cell(row=sheet_a, column=25).value = dados["fator_orc"]
        sheet.cell(row=sheet_a, column=26).value = dados["valor_orc"]

        wb.save(planilha_cadastro_path)
        return True, sheet_a - 16

    except Exception as e:
        return False, str(e)
